# entrar no site do pje "https://pje-consulta-publica.tjmg.jus.br/"
# digitar nº oab e selecionar estado
# clicar em pesquisar
# entrar em cada um dos processos
# extrair o nº do processo e a dta de distribuição
# extrair todas as últimas movimentações
# guardar tudo em uma aba da planilha excel por processo

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select

import openpyxl

from time import sleep

class Scraping:
    def __init__(self):
        self.oab = input("Digite o número da OAB\n")
        self.state = input("Digite a sigla do estado\n").upper()
        
    def Start(self):
        self.open_website()
        self.find_lawyer()
        self.get_info()

    def open_website(self):
        self.driver = webdriver.Chrome()
        link = "https://pje-consulta-publica.tjmg.jus.br/"
        self.driver.get(link)
        sleep(10)

    def find_lawyer(self):
        oab_field = self.driver.find_element(By.XPATH, '//input[@id="fPP:Decoration:numeroOAB"]')
        oab_field.send_keys(self.oab)

        dropdown_state = self.driver.find_element(By.XPATH, '//select[@id="fPP:Decoration:estadoComboOAB"]')
        state_options = Select(dropdown_state)
        state_options.select_by_visible_text(self.state)

        btn_search = self.driver.find_element(By.XPATH, '//input[@id="fPP:searchProcessos"]')
        btn_search.click()
        sleep(15)
        
    def get_info(self):
        processos = self.driver.find_elements(By.XPATH, '//a[@class="btn btn-default btn-sm"]')
        
        for link in processos:
            link.click()
            sleep(8)
            # verificar janelas disponiveis
            windows = self.driver.window_handles
            self.driver.switch_to.window(windows[-1])
            self.driver.set_window_size(1440,802)

            self.num_processo = self.driver.find_elements(By.XPATH, '//div[@class="col-sm-12 "]')[0].text
            print('Nº Processo:', self.num_processo)

            self.data_distribuicao = self.driver.find_elements(By.XPATH, '//div[@class="value col-sm-12 "]')[1].text
            print('Data de Distribuição:', self.data_distribuicao)

            self.classe_judicial = self.driver.find_elements(By.XPATH, '//div[@class="value col-sm-12 "]')[2].text
            print('Classe Judicial:', self.classe_judicial)

            self.assunto = self.driver.find_elements(By.XPATH, '//div[@class="value col-sm-12 "]')[3].text
            print('Assunto:', self.assunto)

            self.jurisdicao = self.driver.find_elements(By.XPATH, '//div[@class="value col-sm-12 "]')[4].text
            print('Jurisdição:', self.jurisdicao)

            self.orgao_julgador = self.driver.find_elements(By.XPATH, '//div[@class="value col-sm-12 "]')[6].text
            print('Orgão Julgador:', self.orgao_julgador)

            movimentacoes_max_page = int(self.driver.find_element(By.XPATH,'//td[@class="rich-inslider-right-num "]').text)
            print("Movimentações, max-page:", movimentacoes_max_page)

            self.lista_movimentacoes = []
            for i in range(movimentacoes_max_page):
                print("Movimentações, current-page:", i+1)

                movimentacoes = self.driver.find_elements(By.XPATH,'//div[@id="j_id132:processoEventoPanel_body"]//tr[contains(@class,"rich-table-row")]//td//div//div//span')
                
                for movimentacao in movimentacoes:
                    self.lista_movimentacoes.append(movimentacao.text)

                if (i+1<movimentacoes_max_page):
                    sleep(3)
                    page_field = self.driver.find_element(By.XPATH, '//input[@name="j_id132:j_id529:j_id530"]')
                    page_field.click()
                    page_field.send_keys(Keys.CONTROL, 'a')
                    page_field.send_keys(i+2)
                    sleep(3)
                else:
                    print("Nº de itens lista_movimentacoes:", len(self.lista_movimentacoes))
                    
            self.insert_info_excel()
            

    def insert_info_excel(self):

        workbook = openpyxl.load_workbook('processos.xlsx')

        try:
            # acessar página do processo (página existente)
            pagina_processo = workbook[self.num_processo]
            # criar nome das colunas
            pagina_processo['A1'].value = 'Número Processo'
            pagina_processo['B1'].value = 'Data da Distribuição'
            pagina_processo['C1'].value = 'Classe Judicial'
            pagina_processo['D1'].value = 'Assunto'
            pagina_processo['E1'].value = 'Jurisdição'
            pagina_processo['F1'].value = 'Órgão Julgador'
            pagina_processo['G1'].value = 'Movimentações do Processo'
            # adicionar informações
            pagina_processo['A2'].value = self.num_processo
            pagina_processo['B2'].value = self.data_distribuicao
            pagina_processo['C2'].value = self.classe_judicial
            pagina_processo['D2'].value = self.assunto
            pagina_processo['E2'].value = self.jurisdicao
            pagina_processo['F2'].value = self.orgao_julgador
            for index, row in enumerate(pagina_processo.iter_rows(min_row=2,max_row=len(self.lista_movimentacoes)+1,min_col=7,max_col=7)):
                for cell in row:
                    cell.value = self.lista_movimentacoes[index]
            # salvar planilha
            workbook.save('processos.xlsx')
            self.driver.close()
            self.driver.switch_to.window(self.driver.window_handles[0])
        except Exception as error:
            # criar página (sheet) para inserir informações
            workbook.create_sheet(self.num_processo)
            # acessar página do processo (página existente)
            pagina_processo = workbook[self.num_processo]
            # criar nome das colunas
            pagina_processo['A1'].value = 'Número Processo'
            pagina_processo['B1'].value = 'Data da Distribuição'
            pagina_processo['C1'].value = 'Classe Judicial'
            pagina_processo['D1'].value = 'Assunto'
            pagina_processo['E1'].value = 'Jurisdição'
            pagina_processo['F1'].value = 'Órgão Julgador'
            pagina_processo['G1'].value = 'Movimentações do Processo'
            # adicionar informações
            pagina_processo['A2'].value = self.num_processo
            pagina_processo['B2'].value = self.data_distribuicao
            pagina_processo['C2'].value = self.classe_judicial
            pagina_processo['D2'].value = self.assunto
            pagina_processo['E2'].value = self.jurisdicao
            pagina_processo['F2'].value = self.orgao_julgador
            for index, row in enumerate(pagina_processo.iter_rows(min_row=2,max_row=len(self.lista_movimentacoes)+1,min_col=7,max_col=7)):
                for cell in row:
                    cell.value = self.lista_movimentacoes[index]
            # salvar planilha
            workbook.save('processos.xlsx')
            self.driver.close()
            self.driver.switch_to.window(self.driver.window_handles[0])
        
            